import streamlit as st
import pandas as pd
from PIL import Image
import os
import sys

# PyInstaller için dosya yolu düzeltici
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# ------------------ Streamlit Uygulaması ------------------

excel_path = resource_path("MOBYA JULY PRODUCT LIST ERCAN.xlsx")
df = pd.read_excel(excel_path)

st.title("Mobilya Ürün Sorgulama Uygulaması")

if "sepet" not in st.session_state:
    st.session_state.sepet = []

# 🔽 Seri numarası seçimi
seri = st.selectbox("Seri Numarası Seçin", options=[""] + df["Serial No."].astype(str).tolist())

if seri:
    if seri in df["Serial No."].astype(str).values:
        urun = df[df["Serial No."].astype(str) == seri].iloc[0]
        st.write(f"**Kategori:** {urun['Main Category']}")
        st.write(f"**Tür:** {urun['Type']}")
        st.write(f"**Birim Fiyatı:** {urun['Unit Price']} $")

        # Görsel yükleme
        image_path = resource_path(f"images/{seri}.png")
        if os.path.exists(image_path):
            st.image(Image.open(image_path), width=125)
        else:
            st.warning("Resim bulunamadı.")

        adet = st.number_input("Adet", min_value=1, value=1, step=1)

        if st.button("Sepete Ekle"):
            st.success(f"{adet} adet {seri} kodlu ürün sepete eklendi.")
            st.session_state.sepet.append({
                "Seri No": seri,
                "Tür": urun["Type"],
                "Adet": adet,
                "Birim Fiyat": urun["Unit Price"],
                "Toplam Fiyat": round(adet * urun["Unit Price"], 4),
                "Görsel": image_path if os.path.exists(image_path) else None
            })
    else:
        st.error("Bu seri numarası bulunamadı.")

# 🛒 Sepet
if st.session_state.sepet:
    st.subheader("🛒 Sepetiniz")
    toplam = 0
    for item in st.session_state.sepet:
        cols = st.columns([1.2, 2, 1, 1.2, 1.5, 1.2])
        if item["Görsel"]:
            cols[0].image(item["Görsel"], width=125)
        else:
            cols[0].write("—")
        cols[1].write(f"**{item['Seri No']}**")
        cols[2].write(item["Tür"])
        cols[3].write(f"x{item['Adet']}")
        cols[4].write(f"{item['Birim Fiyat']:.4f} $")
        cols[5].write(f"{item['Toplam Fiyat']:.4f} $")
        toplam += item["Toplam Fiyat"]

    st.markdown("### Genel Toplam")
    st.success(f"Toplam Fiyat: {round(toplam, 2)} $")

# ------------------ Excel'e Aktarma ------------------
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

def sepeti_duzenli_excel_aktar(sepet, dosya_adi="duzenli_sepet.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sepet"

    basliklar = ["Görsel", "Seri No", "Ürün Tipi", "Adet", "Birim Fiyat ($)", "Toplam Fiyat ($)"]
    ws.append(basliklar)

    for col_num, baslik in enumerate(basliklar, 1):
        hucre = ws.cell(row=1, column=col_num)
        hucre.font = Font(bold=True)
        hucre.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_num)].width = 18

    for i, item in enumerate(sepet, start=2):
        ws.cell(row=i, column=2, value=item["Seri No"])
        ws.cell(row=i, column=3, value=item["Tür"])
        ws.cell(row=i, column=4, value=item["Adet"])
        ws.cell(row=i, column=5, value=item["Birim Fiyat"])
        ws.cell(row=i, column=6, value=item["Toplam Fiyat"])

        for col in range(2, 7):
            ws.cell(row=i, column=col).alignment = Alignment(horizontal="center", vertical="center")

        if item["Görsel"] and os.path.exists(item["Görsel"]):
            img = PILImage.open(item["Görsel"])
            img.thumbnail((125, 125))
            temp_img_path = f"temp_{i}.png"
            img.save(temp_img_path)
            xl_img = XLImage(temp_img_path)
            xl_img.width, xl_img.height = 125, 125
            ws.row_dimensions[i].height = 100
            ws.add_image(xl_img, f"A{i}")

    toplam_satir = len(sepet) + 2
    ws.merge_cells(start_row=toplam_satir, start_column=1, end_row=toplam_satir, end_column=4)
    ws.cell(row=toplam_satir, column=1, value="Genel Toplam:").font = Font(bold=True)
    ws.cell(row=toplam_satir, column=6, value=sum([item["Toplam Fiyat"] for item in sepet])).font = Font(bold=True)

    wb.save(dosya_adi)

    for i in range(2, 2 + len(sepet)):
        temp_path = f"temp_{i}.png"
        if os.path.exists(temp_path):
            os.remove(temp_path)

    return dosya_adi

if st.button("Excel Olarak İndir"):
    dosya = sepeti_duzenli_excel_aktar(st.session_state.sepet)
    with open(dosya, "rb") as f:
        st.download_button("Excel'i İndir", f, file_name="sepet.xlsx")
